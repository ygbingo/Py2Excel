"""
句式统计：
command: python pattern_count.py debug_result.xlsx
debug_result.xlsx: query和对应slot表，其中query放在第2列，对应的slots放在第3列
output: pattern_result.xlsx
备注：最好把不同意图分表后处理
"""
from collections import defaultdict
from openpyxl import load_workbook

def get_pattern_count(frn):
    base = frn[:-5]
    fwn2 = base + '_slots_count.txt'
    dq = defaultdict(int)
    ds = defaultdict(int)
    domains = []
    domain = ''
    lastDomain = ''
    result_row = 2
    with open(fwn2, 'w') as fw2:
        query_workbook = load_workbook(frn)  # 等待提取文件
        query_sheetnames = query_workbook.get_sheet_names()  # 获得表单名字
        result_workbook = load_workbook('pattern_result.xlsx')
        result_sheetnames = result_workbook.get_sheet_names()
        slots_workbook = load_workbook('slots_result.xlsx')
        slots_sheetnames = slots_workbook.get_sheet_names()
        for query_sheetname in query_sheetnames:
            query_sheet = query_workbook.get_sheet_by_name(query_sheetname)
            result_sheet = result_workbook.get_sheet_by_name(result_sheetnames[0])
            result_sheet.cell(row=1, column=1).value = '意图编号'
            result_sheet.cell(row=1, column=2).value = '句式'
            result_sheet.cell(row=1, column=3).value = '频率'
            slots_sheet = slots_workbook.get_sheet_by_name(slots_sheetnames[0])
            slots_sheet.cell(row=1, column=1).value = 'slotName'
            slots_sheet.cell(row=1, column=2).value = 'slot'
            slots_sheet.cell(row=1, column=3).value = '频率'
            for query_row in range(2, query_sheet.max_row + 1):
                domain = query_sheet.cell(row=query_row, column=1).value
                if domain != lastDomain:
                    domains.append(lastDomain)
                    for k, v in sorted(dq.items(), key=lambda x: x[1], reverse=True):
                        result_sheet.cell(row=result_row, column=1).value = lastDomain
                        result_sheet.cell(row=result_row, column=2).value = k
                        result_sheet.cell(row=result_row, column=3).value = v
                        result_row += 1
                    # for k, v in sorted(ds.items(), key=lambda x: x[1], reverse=True):
                    #     fw2.write("%s\t%d\n" % (k, v))
                    dq.clear()
                    # ds.clear()
                    lastDomain = domain
                query = query_sheet.cell(row=query_row, column=2).value
                slots = query_sheet.cell(row=query_row, column=3).value
                slots = str(slots).split(';')
                # print(slots)
                if slots[0] == 'None':
                    dq[query] += 1
                    continue
                for ii in range(0, len(slots)):
                    slot = slots[ii].split(':')[1].split('/')[0]
                    slot_name = slots[ii].split(':')[0]
                    ds[slot_name+':'+slot] += 1
                    query = query.replace(slots[ii].split(':')[1].split('/')[0], '#' + slot_name + '#')
                dq[query] += 1

        for k, v in sorted(dq.items(), key=lambda x: x[1], reverse=True):
            result_sheet.cell(row=result_row, column=1).value = lastDomain
            result_sheet.cell(row=result_row, column=2).value = k
            result_sheet.cell(row=result_row, column=3).value = v
            result_row += 1
        for k, v in sorted(ds.items(), key=lambda x: x[1], reverse=True):
            slots = k.split(':')
            slotName = slots[0]
            slot = slots[1]
            fw2.write("%s\t%s\t%d\n" % (slotName, slot, v))

        result_workbook.save("pattern_result.xlsx")  # 输出文件

if '__main__' == __name__:
    import sys
    frn = sys.argv[1]
    get_pattern_count(frn)

