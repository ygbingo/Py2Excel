from openpyxl import load_workbook

query_row = 2
query_workbook = load_workbook('fm_core_set.xlsx')
query_sheetnames = query_workbook.get_sheet_names()
query_sheet = query_workbook.get_sheet_by_name(query_sheetnames[0])

slots_count_file = open('fm_query_slots_count.txt')
slots = slots_count_file.readlines()
# print(slotsName)

pattern_workbook = load_workbook('fm_pattern_result.xlsx')
pattern_sheetnames = pattern_workbook.get_sheet_names()
pattern_sheet = pattern_workbook.get_sheet_by_name(pattern_sheetnames[0])

iter = 1
lastDomain = ''
for pattern_row in range(2, pattern_sheet.max_row + 1):
    domain = pattern_sheet.cell(row=pattern_row, column=1).value # 意图编号
    if domain == lastDomain:
        iter += 1
    else:
        iter = 1
        lastDomain = domain
    if iter > 10:
        continue
    pattern = pattern_sheet.cell(row=pattern_row, column=2).value # 句式
    if 'fm' not in domain:
        continue
    if pattern:
        if '-' in pattern:
            continue
        if '#' in pattern:
            slotHead = pattern.index('#')
            slotTail = pattern[slotHead + 1:].index('#') + slotHead + 1
            slotName = pattern[slotHead+1:slotTail]
            for ii in range(0, len(slots)):
                slotsName = slots[ii].split('\t')[0]
                if slotsName == slotName:
                    query = pattern.replace('#'+slotName+'#', slots[ii].split('\t')[1])
                    query_sheet.cell(row=query_row, column=1).value = domain
                    query_sheet.cell(row=query_row, column=2).value = query
                    query_sheet.cell(row=query_row, column=3).value = slotName + ':' + slots[ii].split('\t')[1]
                    query_row += 1
        else:
            query_sheet.cell(row=query_row, column=1).value = domain
            query_sheet.cell(row=query_row, column=2).value = pattern
            query_row += 1

query_workbook.save('fm_core_set.xlsx')



        # for query_sheetname in query_sheetnames:
        #     query_sheet = query_workbook.get_sheet_by_name(query_sheetname)
        #     result_sheet = result_workbook.get_sheet_by_name(result_sheetnames[0])
        #     result_sheet.cell(row=1, column=1).value = '意图编号'
        #     result_sheet.cell(row=1, column=2).value = '句式'
        #     result_sheet.cell(row=1, column=3).value = '频率'
        #     slots_sheet = slots_workbook.get_sheet_by_name(slots_sheetnames[0])
        #     slots_sheet.cell(row=1, column=1).value = 'slotName'
        #     slots_sheet.cell(row=1, column=2).value = 'slot'
        #     slots_sheet.cell(row=1, column=3).value = '频率'
        #     for query_row in range(2, query_sheet.max_row + 1):
        #         domain = query_sheet.cell(row=query_row, column=1).value
        #         if domain != lastDomain:
        #             domains.append(lastDomain)
        #             for k, v in sorted(dq.items(), key=lambda x: x[1], reverse=True):
        #                 result_sheet.cell(row=result_row, column=1).value = lastDomain
        #                 result_sheet.cell(row=result_row, column=2).value = k
        #                 result_sheet.cell(row=result_row, column=3).value = v
        #                 result_row += 1
        #             # for k, v in sorted(ds.items(), key=lambda x: x[1], reverse=True):
        #             #     fw2.write("%s\t%d\n" % (k, v))
        #             dq.clear()
        #             # ds.clear()
        #             lastDomain = domain
        #         query = query_sheet.cell(row=query_row, column=2).value
        #         slots = query_sheet.cell(row=query_row, column=3).value
        #         slots = str(slots).split(';')
        #         # print(slots)
        #         if slots[0] == 'None':
        #             dq[query] += 1
        #             continue
        #         for ii in range(0, len(slots)):
        #             slot = slots[ii].split(':')[1].split('/')[0]
        #             slot_name = slots[ii].split(':')[0]
        #             ds[slot_name+':'+slot] += 1
        #             query = query.replace(slots[ii].split(':')[1].split('/')[0], '#' + slot_name + '#')
        #         dq[query] += 1
        #
        # for k, v in sorted(dq.items(), key=lambda x: x[1], reverse=True):
        #     result_sheet.cell(row=result_row, column=1).value = lastDomain
        #     result_sheet.cell(row=result_row, column=2).value = k
        #     result_sheet.cell(row=result_row, column=3).value = v
        #     result_row += 1
        # for k, v in sorted(ds.items(), key=lambda x: x[1], reverse=True):
        #     slots = k.split(':')
        #     slotName = slots[0]
        #     slot = slots[1]
        #     fw2.write("%s\t%s\t%d\n" % (slotName, slot, v))
        #
        # result_workbook.save("pattern_result.xlsx")  # 输出文件