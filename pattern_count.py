"""
句式统计：
command: python pattern_count.py debug_result.xlsx
debug_result.xlsx: query和对应slot表，其中query放在第2列，对应的slots放在第3列
output: debug_result_pattern_count.txt
备注：最好把不同意图分表后处理
"""
from collections import defaultdict
from openpyxl import load_workbook

def get_pattern_count(frn):
    base = frn[:-5]
    fwn = base+'_pattern_count.txt'
    fwn2 = base + '_slots_count.txt'
    dq = defaultdict(int)
    ds = defaultdict(int)
    with open(fwn, 'w') as fw, open(fwn2, 'w') as fw2:
        query_workbook = load_workbook(frn)  # 等待提取文件
        query_sheetnames = query_workbook.get_sheet_names()  # 获得表单名字
        for query_sheetname in query_sheetnames:
            query_sheet = query_workbook.get_sheet_by_name(query_sheetname)
            for query_row in range(2, query_sheet.max_row + 1):
                query = query_sheet.cell(row=query_row, column=2).value
                slots = query_sheet.cell(row=query_row, column=3).value
                slots = str(slots).split(';')
                # print(slots)
                if slots[0] == 'None':
                    dq[query] += 1
                    continue
                for ii in range(0, len(slots)):
                    slot = slots[ii].split(':')[1].split('/')[0]
                    slot_name = slots[ii].split(':')[0]
                    ds[slot] += 1
                    query = query.replace(slots[ii].split(':')[1].split('/')[0], '#'+slot_name+'#')
                dq[query] += 1

        for k, v in sorted(dq.items(), key=lambda x: x[1], reverse=True):
            fw.write("%s\t%d\n" % (k, v))
        for k, v in sorted(ds.items(), key=lambda x: x[1], reverse=True):
            fw2.write("%s\t%d\n" % (k, v))

if '__main__' == __name__:
    import sys
    frn = sys.argv[1]
    get_pattern_count(frn)

