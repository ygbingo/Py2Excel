"""
从query中提取slots值

car_control_slots.xlsx: 存放词槽(根据权重，优先级从高致低排列)
    sheet：为slot名
    value列：slot内容
    slots：slot值

car_control_query.xlsx: 待提取文件
    意图编号*：必填
    query*：必填
    slots：选填(可能被覆盖)
    备注：选填

car_control_query.xls: 输出文件(如果不存在，会自动创建)
    意图编号*
    query*
    slots
    备注
"""
from openpyxl import load_workbook

slot_workbook = load_workbook(u"fm_slots.xlsx")  # 词槽整合文件
query_workbook = load_workbook(u"电台.xlsx")  # 等待提取文件

query_sheetnames = query_workbook.get_sheet_names()  # 获得表单名字
slot_sheetnames = slot_workbook.get_sheet_names()

# query_sheet = query_workbook.get_sheet_by_name(query_sheetnames[0]) # 存放内容的表格，默认为"Sheet1"

for query_sheetname in query_sheetnames:
    query_sheet = query_workbook.get_sheet_by_name(query_sheetname)

    for query_row in range(2, query_sheet.max_row + 1):
        print(query_row)
        row_slots = ""
        row_flag = 0
        val = query_sheet.cell(row=query_row, column=2).value
        """
        导出的测试集在query列会带一些标记过的slot(可能slot有问题)
        需要删除这些[***]
        """
        if val is None: continue
        if '[' in val:
            valTail = val.index('[')
            val = val[:valTail]
            query_sheet.cell(row=query_row, column=2).value = val
        for slot_sheetname in slot_sheetnames:
            # print(slot_sheetname)
            slot_sheet = slot_workbook.get_sheet_by_name(slot_sheetname)
            for slow_row in range(2, slot_sheet.max_row + 1):
                str_slot = str(slot_sheet.cell(row=slow_row, column=1).value)
                str_slot.replace(" ","")
                str_slot.replace("\t","")
                str_slot.replace("\r","")
                str_slot.replace("\n","")
                int_slot = slot_sheet.cell(row=slow_row, column=2).value
                if str_slot in val:
                    """
                    分号在每个slots的后面:  
                    row_slots += slot_sheetname + ":" + str_slot + "/" + str(int_slot) + ";"
                    break
                    """
                    if row_flag == 0:
                        row_slots += slot_sheetname + ":" + str_slot + "/" + str(int_slot)
                        row_flag = 1
                        break
                    else:
                        row_slots += ";" + slot_sheetname + ":" + str_slot + "/" + str(int_slot)
                        break
                    tail = val.find(str_slot)
                    val = val[:tail]+val[tail+len(str_slot):]

        query_sheet.cell(row=query_row, column=3).value = row_slots

query_workbook.save("fm_query.xlsx")  # 输出文件
