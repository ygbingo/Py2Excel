from openpyxl import load_workbook

query_workbook2 = load_workbook('pattern_result.xlsx')
query_sheetnames2 = query_workbook2.get_sheet_names()

print(query_sheetnames2[0])