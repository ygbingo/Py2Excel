from openpyxl import load_workbook

f2 = open('res_poitype','r+')

query_workbook = load_workbook(u"location.xlsx")  # 等待提取文件

query_sheetnames = query_workbook.get_sheet_names()  # 获得表单名字

# query_sheet = query_workbook.get_sheet_by_name(query_sheetnames[0]) # 存放内容的表格，默认为"Sheet1"

for query_sheetname in query_sheetnames:
    query_sheet = query_workbook.get_sheet_by_name(query_sheetname)

    for query_row in range(2, query_sheet.max_row + 1):
        row_slots = ""
        row_flag = 0
        val = query_sheet.cell(row=query_row, column=2).value
        """
        导出的测试集在query列会带一些标记过的slot(可能slot有问题)
        需要删除这些[***]
        """
        findStr = 'poi_type'
        if findStr in val:
            valHead = val.index(findStr) + len(findStr) + 1
            val = val[valHead:-1]
            if ';' in val:
                valTail = val.index(';')
                val = val[:valTail]
                f2.write('|' + val + '\n')
            else:
                f2.write('|' + val + '\n')

f2.close()